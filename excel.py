'''import pandas as pd
import xlrd
from xlutils.copy import copy
df1 = pd.read_excel("C:\\Users\\杨天佑\\Desktop\\资料\\data.xls", encoding = 'utf-8')
df2 = pd.read_excel("C:\\Users\\杨天佑\\Desktop\\资料\\机动车.xlsx", encoding = 'utf-8')
'''#print(df1)
#print(df2)'''
'''c=df1.merge(df2[['纳税人名称','机动']],how='outer',on='纳税人名称')
# c=pd.merge(a,b,how='right')
# print(c.head())
c.to_excel('C:\\Users\\杨天佑\\Desktop\\资料\\test2.xls', encoding = 'utf-8')'''
#from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.writer.excel import ExcelWriter
wb = load_workbook ('C:\\Users\\杨天佑\\Desktop\\test\\test1.xlsx')
ws = wb['Sheet1']
sheet = wb.active
a = sheet.max_row
q=[]
i=0
while i < a:
    i = i+1
    print(ws.cell (i,2).value)
    q.insert(i, ws.cell (i,2).value)
    print(q)

wbtest = load_workbook ('C:\\Users\\杨天佑\\Desktop\\test\\test2.xlsx')
ws = wbtest['Sheet1']
sheet = wb.active

while i < a:
    ws.cell (i+1,1).value = q[i]
    i = i+1
wbtest.save ('C:\\Users\\杨天佑\\Desktop\\test\\test2.xlsx')
